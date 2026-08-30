from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from cmc_bbdm.damage_response.sources import (
    SourceError,
    load_official_inventory,
    load_spatial_pairs,
    read_lvi_observations,
    read_published_peaks,
    read_specimen_sizes,
)


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_inventory(path: Path, records: list[dict[str, object]]) -> str:
    payload = "".join(
        json.dumps(record, sort_keys=True) + "\n" for record in records
    ).encode("ascii")
    path.write_bytes(payload)
    return hashlib.sha256(payload).hexdigest()


def _inventory_record(*, file_id: str, filename: str, folder: str) -> dict[str, object]:
    return {
        "dataset_id": "8scdmfdcfb",
        "file_id": file_id,
        "filename": filename,
        "folder": folder,
        "relative_path": f"8scdmfdcfb/v3/{folder}/{filename}",
        "sha256": hashlib.sha256(filename.encode("ascii")).hexdigest(),
        "size": 123,
        "url": f"https://data.mendeley.com/public-files/{file_id}",
        "version": 3,
    }


def test_official_inventory_is_hash_bound_and_schema_checked(tmp_path: Path) -> None:
    path = tmp_path / "inventory.jsonl"
    records = [
        _inventory_record(
            file_id="raw-1",
            filename="c8-2 0000.CSV",
            folder="4_Compression after impact testing raw data",
        ),
        _inventory_record(
            file_id="image-1",
            filename="c8-2_front.jpg",
            folder="3_Specimen image",
        ),
    ]
    expected_sha = _write_inventory(path, records)

    loaded = load_official_inventory(
        path,
        expected_sha256=expected_sha,
        expected_folder_counts={
            "3_Specimen image": 1,
            "4_Compression after impact testing raw data": 1,
        },
    )

    assert [record.file_id for record in loaded] == ["image-1", "raw-1"]
    assert all("/tmp/" not in repr(record) for record in loaded)


def test_official_inventory_rejects_hash_drift(tmp_path: Path) -> None:
    path = tmp_path / "inventory.jsonl"
    _write_inventory(
        path,
        [
            _inventory_record(
                file_id="raw-1",
                filename="c8-2 0000.CSV",
                folder="4_Compression after impact testing raw data",
            )
        ],
    )

    with pytest.raises(SourceError, match="inventory SHA-256"):
        load_official_inventory(
            path,
            expected_sha256="0" * 64,
            expected_folder_counts={
                "4_Compression after impact testing raw data": 1
            },
        )


def test_official_inventory_rejects_absolute_serialized_path(tmp_path: Path) -> None:
    path = tmp_path / "inventory.jsonl"
    record = _inventory_record(
        file_id="raw-1",
        filename="c8-2 0000.CSV",
        folder="4_Compression after impact testing raw data",
    )
    record["relative_path"] = "/private/c8-2.CSV"
    expected_sha = _write_inventory(path, [record])

    with pytest.raises(SourceError, match="relative path"):
        load_official_inventory(
            path,
            expected_sha256=expected_sha,
            expected_folder_counts={
                "4_Compression after impact testing raw data": 1
            },
        )


def _save_size_workbook(path: Path) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Specimen size"
    sheet["B2"] = "Specimen No."
    sheet["C2"] = "Specimen size[mm]"
    sheet["C3"] = "Height"
    sheet["D3"] = "Width"
    sheet["E3"] = "Thickness"
    sheet.append((None, None, None, None, None))
    sheet["B4"] = "c8-2"
    sheet["C4"] = 80.0
    sheet["D4"] = 50.0
    sheet["E4"] = 2.0
    workbook.save(path)
    workbook.close()
    return _sha(path)


def test_specimen_size_reader_preserves_measured_dimensions(tmp_path: Path) -> None:
    path = tmp_path / "sizes.xlsx"
    expected_sha = _save_size_workbook(path)

    records = read_specimen_sizes(path, expected_sha256=expected_sha)

    assert records["c8-2"].height_mm == 80.0
    assert records["c8-2"].width_mm == 50.0
    assert records["c8-2"].thickness_mm == 2.0


def _save_peak_workbook(path: Path) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CAI strength"
    sheet["B2"] = "Specimen No."
    sheet["C2"] = "Compression after impact strength"
    sheet["C3"] = "[Mpa]"
    sheet["D3"] = "[%]"
    sheet["B4"] = "c8-2"
    sheet["C4"] = 250.1234
    sheet["C4"].number_format = "0.00"
    sheet["D4"] = 0.8
    workbook.save(path)
    workbook.close()
    return _sha(path)


def test_published_peak_reader_uses_cell_display_precision(tmp_path: Path) -> None:
    path = tmp_path / "peaks.xlsx"
    expected_sha = _save_peak_workbook(path)

    records = read_published_peaks(path, expected_sha256=expected_sha)

    assert records["c8-2"].value_mpa == 250.1234
    assert records["c8-2"].decimal_places == 2


def _save_lvi_workbook(path: Path) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LVI condition"
    sheet["B2"] = "Specimen No."
    sheet["C3"] = "Layup"
    sheet["D3"] = "Impactor shape"
    sheet["E3"] = "Impact energy"
    sheet["G3"] = "Projected delamination area"
    sheet["H3"] = "Dent depth"
    sheet["I3"] = "Is included"
    sheet["E4"] = "[J/mm]"
    sheet["F4"] = "[J]"
    sheet["B5"] = "c8-2"
    sheet["C5"] = "c8"
    sheet["D5"] = "hemia"
    sheet["E5"] = 3.3
    sheet["F5"] = 6.6
    sheet["G5"] = 120.0
    sheet["H5"] = 0.5
    sheet["I5"] = "yes"
    workbook.save(path)
    workbook.close()
    return _sha(path)


def test_lvi_reader_marks_numeric_pre_cai_damage_observation(tmp_path: Path) -> None:
    path = tmp_path / "lvi.xlsx"
    expected_sha = _save_lvi_workbook(path)

    records = read_lvi_observations(path, expected_sha256=expected_sha)

    assert records["c8-2"].has_numeric_damage_observation is True
    assert records["c8-2"].is_intact is False


def test_spatial_manifest_returns_identities_without_private_paths(
    tmp_path: Path,
) -> None:
    path = tmp_path / "paired.csv"
    fields = (
        "sample_id",
        "dataset_id",
        "source_sha256",
        "target_sha256",
        "surface_profile_before_path",
        "surface_profile_after_path",
    )
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerow(
            {
                "sample_id": "c8-2",
                "dataset_id": "74t7kcdgkr",
                "source_sha256": "a" * 64,
                "target_sha256": "b" * 64,
                "surface_profile_before_path": "/private/before.csv",
                "surface_profile_after_path": "/private/after.csv",
            }
        )

    records = load_spatial_pairs(path, expected_sha256=_sha(path))

    assert records[0].specimen_id == "c8-2"
    assert records[0].domain_id == "74t7kcdgkr"
    assert records[0].cscan_sha256 == "b" * 64
    assert records[0].surface_profile_available is True
    assert "/private/" not in repr(records[0])
