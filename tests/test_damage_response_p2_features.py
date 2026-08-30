from __future__ import annotations

import csv
import hashlib
from pathlib import Path

import numpy as np
import pytest
import yaml
from openpyxl import Workbook

from cmc_bbdm.damage_response.p2_features import (
    FEATURE_CACHE_MEMBERS,
    PHYSICAL_DESCRIPTOR_FIELDS,
    PROFILE_STAT_NAMES,
    PROVENANCE_FIELDS,
    P2FeatureError,
    load_p2_feature_authority,
    serialize_feature_authority_csv,
)
from cmc_bbdm.damage_response.sources import DesignMetadata

ROOT = Path(__file__).resolve().parents[1]
CONFIG = ROOT / "paper_v3/configs/damage_to_failure_response_p2.yaml"
_LOAD_KEYS = (
    "roster",
    "feature_bank_path",
    "feature_bank_sha256",
    "feature_cache_path",
    "feature_cache_sha256",
    "physical_descriptors_path",
    "physical_descriptors_sha256",
    "provenance_path",
    "provenance_sha256",
    "lvi_workbook_path",
    "lvi_workbook_sha256",
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_csv(path: Path, fields: tuple[str, ...], rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def _write_lvi(path: Path, records: tuple[DesignMetadata, ...]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LVI condition"
    for address, value in {
        "B2": "Specimen No.",
        "C3": "Layup",
        "D3": "Impactor shape",
        "E3": "Impact energy",
        "G3": "Projected delamination area",
        "H3": "Dent depth",
        "I3": "Is included",
        "E4": "[J/mm]",
        "F4": "[J]",
    }.items():
        sheet[address] = value
    for row_index, record in enumerate(records, start=5):
        prefix = "c" if record.laminate_type == "cross_ply" else "q"
        sheet.cell(row_index, 2, record.specimen_id)
        sheet.cell(row_index, 3, f"{prefix}{record.ply_count}")
        sheet.cell(row_index, 4, record.impactor)
        sheet.cell(row_index, 5, 10.0 + row_index)  # E: total J despite E4.
        sheet.cell(row_index, 6, 1.0 + row_index)  # F: J/mm despite F4.
        sheet.cell(row_index, 7, 100.0)
        sheet.cell(row_index, 8, 0.1)
        sheet.cell(row_index, 9, "yes")
    workbook.save(path)


@pytest.fixture
def feature_sources(tmp_path: Path) -> dict[str, object]:
    records = (
        DesignMetadata("s-1", "d1", "cross_ply", 8, "hemia", 50.0, 2.0),
        DesignMetadata("s-2", "d1", "quasi_isotropic", 16, "coni60", 51.0, 2.1),
        DesignMetadata("s-3", "d2", "cross_ply", 24, "flat", 52.0, 2.2),
    )
    ids = np.asarray([record.specimen_id for record in records])
    domains = np.asarray([record.domain_id for record in records])

    feature_bank = tmp_path / "paired_features.npz"
    embeddings = np.arange(3 * 3 * 512, dtype=np.float32).reshape(3, 3, 512)
    np.savez(
        feature_bank,
        dataset_ids=domains,
        encoder_sha256=np.asarray(["a" * 64]),
        features=embeddings,
        specimen_ids=ids,
        state_sha256=np.asarray(["b" * 64]),
        view_names=np.asarray(["FULL", "BILINEAR_50", "BILINEAR_25"]),
    )

    cache = tmp_path / "feature_cache.npz"
    cache_ids = np.asarray(["extra", *ids.tolist()])
    cache_domains = np.asarray(["other", *domains.tolist()])
    profiles = np.arange(4 * 21, dtype=np.float64).reshape(4, 21) / 10.0
    np.savez(
        cache,
        dataset_ids=cache_domains,
        metadata=np.zeros((4, 10), dtype=np.float64),
        profile_map=np.zeros((4, 1024), dtype=np.float32),
        profile_stats=profiles,
        rgb=np.zeros((4, 512), dtype=np.float32),
        sample_ids=cache_ids,
        schema_version=np.asarray([1], dtype=np.int64),
        targets=np.full((4, 6), 987654.0, dtype=np.float64),
    )

    descriptor_rows: list[dict[str, object]] = []
    provenance_rows: list[dict[str, object]] = []
    for index, record in enumerate(records):
        source_hash = f"{index + 1:064x}"
        descriptor = dict.fromkeys(PHYSICAL_DESCRIPTOR_FIELDS, "")
        descriptor.update(
            {
                "sample_id": record.specimen_id,
                "dataset_id": record.domain_id,
                "descriptor_source": "ultrasonic_cscan_measurement",
                "projected_damage_area": 100.0 + index,
                "projected_damage_area_unit": "mm^2",
                "damage_height": 20.0 + index,
                "damage_height_unit": "mm",
                "damage_width": 30.0 + index,
                "damage_width_unit": "mm",
                "measurement_mask_sha256": f"{index + 11:064x}",
                "source_sha256": source_hash,
            }
        )
        for name in PHYSICAL_DESCRIPTOR_FIELDS:
            if name.endswith("_unit") and descriptor[name] == "":
                descriptor[name] = "1"
            elif descriptor[name] == "" and name not in {
                "sample_id",
                "dataset_id",
                "descriptor_source",
                "measurement_mask_sha256",
                "source_sha256",
            }:
                descriptor[name] = "0"
        descriptor_rows.append(descriptor)

        provenance = dict.fromkeys(PROVENANCE_FIELDS, "")
        provenance.update(
            {
                "specimen_uid": record.specimen_id,
                "source_dataset": record.domain_id,
                "source_version": "3",
                "material_family": "CFRP",
                "layup": (
                    f"cross_ply_{record.ply_count}"
                    if record.laminate_type == "cross_ply"
                    else f"quasi_isotropic_{record.ply_count}"
                ),
                "ply_count": str(record.ply_count),
                "impactor": record.impactor,
                "surface_profile_available": "true",
                "cscan_available": "true",
                "strength_available": "true",
                "cscan_source_sha256": source_hash,
                "pairing_status": "complete_surface_cscan_cai",
                "pairing_method": "exact_published_specimen_id",
                "pairing_confidence": "exact",
                "split_group": record.domain_id,
            }
        )
        provenance_rows.append(provenance)

    descriptors = tmp_path / "physical_descriptors.csv"
    provenance = tmp_path / "specimens.csv"
    _write_csv(descriptors, PHYSICAL_DESCRIPTOR_FIELDS, descriptor_rows)
    _write_csv(provenance, PROVENANCE_FIELDS, provenance_rows)
    lvi = tmp_path / "lvi.xlsx"
    _write_lvi(lvi, records)

    return {
        "roster": records,
        "feature_bank_path": feature_bank,
        "feature_bank_sha256": _sha256(feature_bank),
        "feature_cache_path": cache,
        "feature_cache_sha256": _sha256(cache),
        "physical_descriptors_path": descriptors,
        "physical_descriptors_sha256": _sha256(descriptors),
        "provenance_path": provenance,
        "provenance_sha256": _sha256(provenance),
        "lvi_workbook_path": lvi,
        "lvi_workbook_sha256": _sha256(lvi),
        "embeddings": embeddings,
        "profiles": profiles[1:],
    }


def _load(sources: dict[str, object]):
    return load_p2_feature_authority(
        **{name: sources[name] for name in _LOAD_KEYS}
    )


def test_p2_config_freezes_authority_and_low_capacity_protocol() -> None:
    config = yaml.safe_load(CONFIG.read_text(encoding="utf-8"))

    assert config["p1"]["required_status"] == "P1_GO"
    assert config["p1"]["summary_sha256"] == (
        "37da95962395a0915f586820ab03f06d8d859856e8637d975bc302b1d555ebc7"
    )
    assert config["cohort"]["expected_n"] == 276
    assert config["features"]["full_embedding_view"] == "FULL"
    assert config["features"]["deployable_views"] == ["F0", "F1", "F2", "F3", "F4"]
    assert config["features"]["privileged_views"] == ["F5"]
    assert config["evaluation"]["estimator"] == "Ridge"
    assert config["evaluation"]["ridge_alphas"] == [0.1, 1.0, 10.0, 100.0]
    assert config["evaluation"]["tie_order"] == [
        "lower_pca_dimension",
        "larger_ridge_alpha",
    ]
    assert config["bootstrap"]["replicates"] == 100000
    assert config["bootstrap"]["primary_contrast_count"] == 6
    assert config["gate"]["primary_reference"] == "F2"
    assert config["gate"]["primary_candidates"] == ["F3", "F4"]
    assert config["input_boundary"]["neural_model_authorized"] is False
    assert config["input_boundary"]["p3_curve_model_authorized"] is False


def test_loads_exact_target_free_feature_authority(
    feature_sources: dict[str, object],
) -> None:
    authority = _load(feature_sources)

    assert authority.specimen_ids == ("s-1", "s-2", "s-3")
    assert authority.domain_ids == ("d1", "d1", "d2")
    assert authority.surface_profile_stats.shape == (3, len(PROFILE_STAT_NAMES))
    np.testing.assert_array_equal(
        authority.surface_profile_stats, feature_sources["profiles"]
    )
    np.testing.assert_array_equal(
        authority.full_cscan_embedding,
        np.asarray(feature_sources["embeddings"])[:, 0, :],
    )
    np.testing.assert_array_equal(
        authority.scalar_damage,
        np.asarray([[100.0, 20.0, 30.0], [101.0, 21.0, 31.0], [102.0, 22.0, 32.0]]),
    )
    # The registered workbook semantics use original column E as total joules.
    np.testing.assert_array_equal(authority.privileged_total_energy_j, [15.0, 16.0, 17.0])
    assert authority.privileged_impactors == ("hemia", "coni60", "flat")
    assert authority.full_embedding_view == "FULL"


def test_feature_authority_serialization_is_deterministic_and_target_free(
    feature_sources: dict[str, object],
) -> None:
    authority = _load(feature_sources)

    first = serialize_feature_authority_csv(authority)
    second = serialize_feature_authority_csv(authority)
    lowered = first.decode("ascii").lower()

    assert first == second
    assert "987654" not in lowered
    for forbidden in ("target", "strength", "response", "trace", "post_cai"):
        assert forbidden not in lowered
    assert "full_embedding_row_sha256" in lowered
    assert lowered.count("\n") == 4


def test_rejects_source_hash_drift(feature_sources: dict[str, object]) -> None:
    feature_sources["feature_cache_sha256"] = "0" * 64

    with pytest.raises(P2FeatureError, match="SHA-256 mismatch"):
        _load(feature_sources)


def test_rejects_feature_bank_view_drift(feature_sources: dict[str, object]) -> None:
    path = Path(feature_sources["feature_bank_path"])
    with np.load(path, allow_pickle=False) as values:
        payload = {name: values[name] for name in values.files}
    payload["view_names"] = np.asarray(["BILINEAR_50", "FULL", "BILINEAR_25"])
    np.savez(path, **payload)
    feature_sources["feature_bank_sha256"] = _sha256(path)

    with pytest.raises(P2FeatureError, match="view registry"):
        _load(feature_sources)


def test_rejects_unexpected_feature_cache_member(
    feature_sources: dict[str, object],
) -> None:
    path = Path(feature_sources["feature_cache_path"])
    with np.load(path, allow_pickle=False) as values:
        payload = {name: values[name] for name in values.files}
    assert set(payload) == set(FEATURE_CACHE_MEMBERS)
    payload["unexpected"] = np.asarray([1])
    np.savez(path, **payload)
    feature_sources["feature_cache_sha256"] = _sha256(path)

    with pytest.raises(P2FeatureError, match="member registry"):
        _load(feature_sources)


def test_rejects_cross_source_domain_mismatch(
    feature_sources: dict[str, object],
) -> None:
    path = Path(feature_sources["feature_bank_path"])
    with np.load(path, allow_pickle=False) as values:
        payload = {name: values[name] for name in values.files}
    payload["dataset_ids"] = np.asarray(["d1", "wrong", "d2"])
    np.savez(path, **payload)
    feature_sources["feature_bank_sha256"] = _sha256(path)

    with pytest.raises(P2FeatureError, match="domain"):
        _load(feature_sources)


def test_rejects_nonfinite_surface_value(feature_sources: dict[str, object]) -> None:
    path = Path(feature_sources["feature_cache_path"])
    with np.load(path, allow_pickle=False) as values:
        payload = {name: values[name] for name in values.files}
    payload["profile_stats"] = np.asarray(payload["profile_stats"]).copy()
    payload["profile_stats"][1, 0] = np.nan
    np.savez(path, **payload)
    feature_sources["feature_cache_sha256"] = _sha256(path)

    with pytest.raises(P2FeatureError, match="finite"):
        _load(feature_sources)
