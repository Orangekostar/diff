from __future__ import annotations

import csv
import hashlib
import io
import math
import re
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType

import numpy as np
from openpyxl import load_workbook

from cmc_bbdm.damage_response.authority import AuthorityError, snapshot_file
from cmc_bbdm.damage_response.sources import (
    IMPACTOR_CATEGORIES,
    LAMINATE_CATEGORIES,
    DesignMetadata,
)

PROFILE_STAT_NAMES = (
    "minimum_delta",
    "maximum_delta",
    "quantile_01",
    "quantile_05",
    "quantile_25",
    "median_delta",
    "quantile_75",
    "quantile_95",
    "quantile_99",
    "mean_delta",
    "standard_deviation",
    "negative_volume_mean",
    "positive_volume_mean",
    "negative_fraction",
    "positive_fraction",
    "gradient_mean",
    "gradient_standard_deviation",
    "gradient_quantile_95",
    "dent_centroid_y",
    "dent_centroid_x",
    "dent_radial_extent",
)
SCALAR_DAMAGE_NAMES = (
    "projected_damage_area",
    "damage_height",
    "damage_width",
)
SCALAR_DAMAGE_UNITS = ("mm^2", "mm", "mm")
FEATURE_BANK_MEMBERS = (
    "dataset_ids",
    "encoder_sha256",
    "features",
    "specimen_ids",
    "state_sha256",
    "view_names",
)
FEATURE_CACHE_MEMBERS = (
    "dataset_ids",
    "metadata",
    "profile_map",
    "profile_stats",
    "rgb",
    "sample_ids",
    "schema_version",
    "targets",
)
PHYSICAL_DESCRIPTOR_FIELDS = (
    "sample_id",
    "dataset_id",
    "descriptor_source",
    "projected_damage_area",
    "projected_damage_area_unit",
    "damage_height",
    "damage_height_unit",
    "damage_width",
    "damage_width_unit",
    "damage_eccentricity",
    "damage_eccentricity_unit",
    "equivalent_damage_diameter",
    "equivalent_damage_diameter_unit",
    "damage_radial_spread",
    "damage_radial_spread_unit",
    "maximum_feret_diameter",
    "maximum_feret_diameter_unit",
    "minimum_feret_diameter",
    "minimum_feret_diameter_unit",
    "major_axis_length",
    "major_axis_length_unit",
    "minor_axis_length",
    "minor_axis_length_unit",
    "damage_orientation_degrees",
    "damage_orientation_degrees_unit",
    "orientation_cos_2theta",
    "orientation_cos_2theta_unit",
    "orientation_sin_2theta",
    "orientation_sin_2theta_unit",
    "perimeter",
    "perimeter_unit",
    "convex_hull_area",
    "convex_hull_area_unit",
    "solidity",
    "solidity_unit",
    "compactness",
    "compactness_unit",
    "aspect_ratio",
    "aspect_ratio_unit",
    "component_count",
    "component_count_unit",
    "measurement_mask_sha256",
    "source_sha256",
)
PROVENANCE_FIELDS = (
    "specimen_uid",
    "source_dataset",
    "source_version",
    "material_family",
    "layup",
    "ply_count",
    "campaign_id",
    "impact_energy_value",
    "impact_energy_unit",
    "impactor",
    "scan_orientation",
    "surface_rgb_available",
    "surface_profile_available",
    "cscan_available",
    "strength_available",
    "surface_source_file",
    "surface_source_sha256",
    "cscan_source_file",
    "cscan_source_sha256",
    "lvi_source_row",
    "size_source_row",
    "cai_source_row",
    "pairing_status",
    "pairing_method",
    "pairing_confidence",
    "exclusion_reason",
    "split_group",
)

_SHA256_RE = re.compile(r"[0-9a-f]{64}")
_MAX_SOURCE_BYTES = 256 * 1024 * 1024
_FULL_VIEW_NAMES = ("FULL", "BILINEAR_50", "BILINEAR_25")


class P2FeatureError(RuntimeError):
    """Raised when a P2 feature source or cross-source binding drifts."""


def _digest(value: object, *, label: str) -> str:
    if not isinstance(value, str):
        raise P2FeatureError(f"{label} must be a SHA-256")
    result = value.strip().casefold()
    if _SHA256_RE.fullmatch(result) is None:
        raise P2FeatureError(f"{label} must be a SHA-256")
    return result


def _bound_bytes(path: Path, expected_sha256: str, *, label: str) -> bytes:
    expected = _digest(expected_sha256, label=f"{label} expected SHA-256")
    candidate = Path(path)
    try:
        snapshot = snapshot_file(
            candidate,
            max_bytes=_MAX_SOURCE_BYTES,
            logical_source=label,
            relative_path=candidate.name,
        )
    except AuthorityError as error:
        raise P2FeatureError(str(error)) from error
    if snapshot.sha256 != expected:
        raise P2FeatureError(
            f"{label} SHA-256 mismatch: expected {expected}, "
            f"observed {snapshot.sha256}"
        )
    try:
        payload = candidate.read_bytes()
    except OSError as error:
        raise P2FeatureError(f"{label} cannot be read") from error
    if (
        len(payload) != snapshot.size
        or hashlib.sha256(payload).hexdigest() != snapshot.sha256
    ):
        raise P2FeatureError(f"{label} changed during read")
    return payload


def _npz_arrays(
    payload: bytes, *, label: str, expected_members: Sequence[str]
) -> dict[str, np.ndarray]:
    try:
        with np.load(io.BytesIO(payload), allow_pickle=False) as values:
            if set(values.files) != set(expected_members):
                raise P2FeatureError(f"{label} member registry changed")
            arrays = {name: np.asarray(values[name]) for name in expected_members}
    except P2FeatureError:
        raise
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        raise P2FeatureError(f"{label} cannot be decoded") from error
    if any(array.dtype.hasobject for array in arrays.values()):
        raise P2FeatureError(f"{label} contains object arrays")
    return arrays


def _normalized_text_vector(
    value: np.ndarray, *, label: str, require_unique: bool
) -> tuple[str, ...]:
    array = np.asarray(value)
    if array.ndim != 1:
        raise P2FeatureError(f"{label} must be one-dimensional")
    result = tuple(str(item).strip().casefold() for item in array.tolist())
    if any(not item for item in result) or (
        require_unique and len(set(result)) != len(result)
    ):
        raise P2FeatureError(f"{label} values are empty or duplicate")
    return result


def _finite_matrix(
    value: np.ndarray, *, label: str, shape: tuple[int, int]
) -> np.ndarray:
    try:
        result = np.asarray(value, dtype=np.float64)
    except (TypeError, ValueError) as error:
        raise P2FeatureError(f"{label} is not numeric") from error
    if result.shape != shape or not np.all(np.isfinite(result)):
        raise P2FeatureError(f"{label} shape or finite-value contract changed")
    return result


def _csv_rows(
    payload: bytes, *, label: str, fields: Sequence[str], id_field: str
) -> dict[str, dict[str, str]]:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise P2FeatureError(f"{label} is not UTF-8") from error
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames != list(fields):
        raise P2FeatureError(f"{label} header registry changed")
    rows: dict[str, dict[str, str]] = {}
    for line, row in enumerate(reader, start=2):
        if set(row) != set(fields) or None in row:
            raise P2FeatureError(f"{label} row schema changed at line {line}")
        specimen_id = str(row[id_field] or "").strip().casefold()
        if not specimen_id or specimen_id in rows:
            raise P2FeatureError(f"{label} specimen IDs are empty or duplicate")
        rows[specimen_id] = {name: str(row[name] or "").strip() for name in fields}
    if not rows:
        raise P2FeatureError(f"{label} is empty")
    return rows


def _finite_csv(value: str, *, label: str, positive: bool = False) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError) as error:
        raise P2FeatureError(f"{label} is not numeric") from error
    if not math.isfinite(result) or (positive and result <= 0.0):
        raise P2FeatureError(f"{label} is not finite and positive")
    return result


@dataclass(frozen=True)
class _LviCondition:
    laminate_type: str
    ply_count: int
    impactor: str
    total_energy_j: float
    energy_per_thickness_j_per_mm: float


def _lvi_conditions(payload: bytes) -> dict[str, _LviCondition]:
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=False)
    except (OSError, ValueError) as error:
        raise P2FeatureError("LVI workbook cannot be decoded") from error
    try:
        if "LVI condition" not in workbook.sheetnames:
            raise P2FeatureError("LVI condition worksheet is absent")
        sheet = workbook["LVI condition"]
        expected_headers = {
            "B2": "Specimen No.",
            "C3": "Layup",
            "D3": "Impactor shape",
            "E3": "Impact energy",
            "G3": "Projected delamination area",
            "H3": "Dent depth",
            "I3": "Is included",
            "E4": "[J/mm]",
            "F4": "[J]",
        }
        if any(sheet[address].value != value for address, value in expected_headers.items()):
            raise P2FeatureError("LVI workbook header registry changed")
        records: dict[str, _LviCondition] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=5, min_col=2, max_col=9, values_only=True),
            start=5,
        ):
            if row[0] is None:
                continue
            specimen_id = str(row[0]).strip().casefold()
            if not specimen_id or specimen_id in records:
                raise P2FeatureError("LVI specimen IDs are empty or duplicate")
            layup = str(row[1] or "").strip().casefold()
            if layup in {"r0", "r45"}:
                continue
            match = re.fullmatch(r"([cq])(\d+)", layup)
            impactor = str(row[2] or "").strip().casefold()
            if match is None:
                raise P2FeatureError(f"LVI layup is invalid at row {row_number}")
            if impactor == "-" and all(str(value).strip() == "-" for value in row[3:7]):
                continue
            if impactor not in IMPACTOR_CATEGORIES:
                raise P2FeatureError(f"LVI impactor is invalid at row {row_number}")
            # Published values and unit cells are reversed: E is total J; F is J/mm.
            total_energy = _finite_csv(
                str(row[3]), label=f"LVI total energy at row {row_number}", positive=True
            )
            per_thickness = _finite_csv(
                str(row[4]),
                label=f"LVI energy per thickness at row {row_number}",
                positive=True,
            )
            records[specimen_id] = _LviCondition(
                laminate_type=(
                    "cross_ply" if match.group(1) == "c" else "quasi_isotropic"
                ),
                ply_count=int(match.group(2)),
                impactor=impactor,
                total_energy_j=total_energy,
                energy_per_thickness_j_per_mm=per_thickness,
            )
    finally:
        workbook.close()
    if not records:
        raise P2FeatureError("LVI workbook contains no registered conditions")
    return records


def _readonly(value: np.ndarray, *, dtype: np.dtype) -> np.ndarray:
    result = np.asarray(value, dtype=dtype).copy()
    result.setflags(write=False)
    return result


@dataclass(frozen=True)
class P2FeatureAuthority:
    specimen_ids: tuple[str, ...]
    domain_ids: tuple[str, ...]
    laminate_types: tuple[str, ...]
    ply_counts: np.ndarray
    widths_mm: np.ndarray
    thicknesses_mm: np.ndarray
    surface_profile_stats: np.ndarray
    scalar_damage: np.ndarray
    full_cscan_embedding: np.ndarray
    privileged_total_energy_j: np.ndarray
    privileged_impactors: tuple[str, ...]
    full_embedding_view: str
    encoder_sha256: str
    embedding_state_sha256: str
    source_sha256: Mapping[str, str]

    def __post_init__(self) -> None:
        n = len(self.specimen_ids)
        if (
            n == 0
            or len(set(self.specimen_ids)) != n
            or len(self.domain_ids) != n
            or len(self.laminate_types) != n
            or len(self.privileged_impactors) != n
        ):
            raise P2FeatureError("P2 feature authority identity registry is invalid")
        if any(not item for item in (*self.specimen_ids, *self.domain_ids)):
            raise P2FeatureError("P2 feature authority contains empty identities")
        if any(item not in LAMINATE_CATEGORIES for item in self.laminate_types):
            raise P2FeatureError("P2 feature authority laminate category changed")
        if any(item not in IMPACTOR_CATEGORIES for item in self.privileged_impactors):
            raise P2FeatureError("P2 feature authority impactor category changed")
        if self.full_embedding_view != "FULL":
            raise P2FeatureError("P2 full embedding view changed")
        _digest(self.encoder_sha256, label="encoder SHA-256")
        _digest(self.embedding_state_sha256, label="embedding state SHA-256")
        source_hashes = {
            str(name): _digest(digest, label=f"{name} source SHA-256")
            for name, digest in self.source_sha256.items()
        }
        if len(source_hashes) != 5:
            raise P2FeatureError("P2 feature source registry changed")

        ply = _readonly(self.ply_counts, dtype=np.dtype(np.int64))
        width = _readonly(self.widths_mm, dtype=np.dtype(np.float64))
        thickness = _readonly(self.thicknesses_mm, dtype=np.dtype(np.float64))
        profile = _readonly(
            self.surface_profile_stats, dtype=np.dtype(np.float64)
        )
        scalar = _readonly(self.scalar_damage, dtype=np.dtype(np.float64))
        embedding = _readonly(
            self.full_cscan_embedding, dtype=np.dtype(np.float32)
        )
        total_energy = _readonly(
            self.privileged_total_energy_j, dtype=np.dtype(np.float64)
        )
        expected_shapes = {
            "ply_counts": (ply, (n,)),
            "widths_mm": (width, (n,)),
            "thicknesses_mm": (thickness, (n,)),
            "surface_profile_stats": (profile, (n, len(PROFILE_STAT_NAMES))),
            "scalar_damage": (scalar, (n, len(SCALAR_DAMAGE_NAMES))),
            "full_cscan_embedding": (embedding, (n, 512)),
            "privileged_total_energy_j": (total_energy, (n,)),
        }
        for label, (array, shape) in expected_shapes.items():
            if array.shape != shape or not np.all(np.isfinite(array)):
                raise P2FeatureError(f"{label} shape or finite-value contract changed")
        if np.any(ply <= 0) or np.any(width <= 0.0) or np.any(thickness <= 0.0):
            raise P2FeatureError("P2 design values must be positive")
        if np.any(scalar < 0.0) or np.any(total_energy <= 0.0):
            raise P2FeatureError("P2 damage/impact values are outside their range")
        object.__setattr__(self, "ply_counts", ply)
        object.__setattr__(self, "widths_mm", width)
        object.__setattr__(self, "thicknesses_mm", thickness)
        object.__setattr__(self, "surface_profile_stats", profile)
        object.__setattr__(self, "scalar_damage", scalar)
        object.__setattr__(self, "full_cscan_embedding", embedding)
        object.__setattr__(self, "privileged_total_energy_j", total_energy)
        object.__setattr__(self, "source_sha256", MappingProxyType(source_hashes))

    @property
    def full_embedding_row_sha256(self) -> tuple[str, ...]:
        return tuple(
            hashlib.sha256(np.ascontiguousarray(row, dtype="<f4").tobytes()).hexdigest()
            for row in self.full_cscan_embedding
        )


def load_p2_feature_authority(
    *,
    roster: Sequence[DesignMetadata],
    feature_bank_path: Path,
    feature_bank_sha256: str,
    feature_cache_path: Path,
    feature_cache_sha256: str,
    physical_descriptors_path: Path,
    physical_descriptors_sha256: str,
    provenance_path: Path,
    provenance_sha256: str,
    lvi_workbook_path: Path,
    lvi_workbook_sha256: str,
) -> P2FeatureAuthority:
    """Load the direct, hash-bound, target-free P2 feature authority."""

    records = tuple(roster)
    specimen_ids = tuple(record.specimen_id.strip().casefold() for record in records)
    domain_ids = tuple(record.domain_id.strip().casefold() for record in records)
    if (
        not records
        or len(set(specimen_ids)) != len(records)
        or any(not value for value in (*specimen_ids, *domain_ids))
    ):
        raise P2FeatureError("P2 roster identities are empty or duplicate")
    if any(record.specimen_id != specimen_id for record, specimen_id in zip(records, specimen_ids, strict=True)):
        raise P2FeatureError("P2 roster specimen IDs are not canonical")

    bank_payload = _bound_bytes(
        feature_bank_path, feature_bank_sha256, label="P2 feature bank"
    )
    cache_payload = _bound_bytes(
        feature_cache_path, feature_cache_sha256, label="P2 feature cache"
    )
    descriptor_payload = _bound_bytes(
        physical_descriptors_path,
        physical_descriptors_sha256,
        label="P2 physical descriptors",
    )
    provenance_payload = _bound_bytes(
        provenance_path, provenance_sha256, label="P2 specimen provenance"
    )
    lvi_payload = _bound_bytes(
        lvi_workbook_path, lvi_workbook_sha256, label="P2 LVI workbook"
    )

    bank = _npz_arrays(
        bank_payload,
        label="P2 feature bank",
        expected_members=FEATURE_BANK_MEMBERS,
    )
    bank_ids = _normalized_text_vector(
        bank["specimen_ids"], label="feature-bank IDs", require_unique=True
    )
    bank_domains = _normalized_text_vector(
        bank["dataset_ids"], label="feature-bank domains", require_unique=False
    )
    if bank_ids != specimen_ids:
        raise P2FeatureError("feature-bank specimen roster differs")
    if bank_domains != domain_ids:
        raise P2FeatureError("feature-bank domain roster differs")
    view_names = tuple(str(item) for item in np.asarray(bank["view_names"]).tolist())
    if view_names != _FULL_VIEW_NAMES:
        raise P2FeatureError("feature-bank view registry changed")
    features = np.asarray(bank["features"])
    if (
        features.dtype != np.float32
        or features.shape != (len(records), len(_FULL_VIEW_NAMES), 512)
        or not np.all(np.isfinite(features))
    ):
        raise P2FeatureError("feature-bank embedding shape or finite values changed")
    encoder = np.asarray(bank["encoder_sha256"])
    state = np.asarray(bank["state_sha256"])
    if encoder.shape != (1,) or state.shape != (1,):
        raise P2FeatureError("feature-bank hash registry changed")
    encoder_sha256 = _digest(str(encoder[0]), label="feature-bank encoder SHA-256")
    state_sha256 = _digest(str(state[0]), label="feature-bank state SHA-256")

    cache = _npz_arrays(
        cache_payload,
        label="P2 feature cache",
        expected_members=FEATURE_CACHE_MEMBERS,
    )
    cache_ids = _normalized_text_vector(
        cache["sample_ids"], label="feature-cache IDs", require_unique=True
    )
    cache_domains = _normalized_text_vector(
        cache["dataset_ids"], label="feature-cache domains", require_unique=False
    )
    cache_order = {specimen_id: index for index, specimen_id in enumerate(cache_ids)}
    if not set(specimen_ids).issubset(cache_order):
        raise P2FeatureError("feature cache does not cover the P2 roster")
    cache_indices = np.asarray([cache_order[item] for item in specimen_ids], dtype=np.int64)
    if tuple(cache_domains[index] for index in cache_indices) != domain_ids:
        raise P2FeatureError("feature-cache domain roster differs")
    if (
        np.asarray(cache["schema_version"]).shape != (1,)
        or int(np.asarray(cache["schema_version"])[0]) != 1
    ):
        raise P2FeatureError("feature-cache schema version changed")
    profile_all = _finite_matrix(
        cache["profile_stats"],
        label="feature-cache surface profile statistics",
        shape=(len(cache_ids), len(PROFILE_STAT_NAMES)),
    )
    if (
        np.asarray(cache["metadata"]).shape != (len(cache_ids), 10)
        or np.asarray(cache["profile_map"]).shape != (len(cache_ids), 1024)
        or np.asarray(cache["rgb"]).shape != (len(cache_ids), 512)
        or np.asarray(cache["targets"]).shape != (len(cache_ids), 6)
    ):
        raise P2FeatureError("feature-cache nonselected array shapes changed")
    profiles = profile_all[cache_indices]

    descriptor_rows = _csv_rows(
        descriptor_payload,
        label="P2 physical descriptors",
        fields=PHYSICAL_DESCRIPTOR_FIELDS,
        id_field="sample_id",
    )
    provenance_rows = _csv_rows(
        provenance_payload,
        label="P2 specimen provenance",
        fields=PROVENANCE_FIELDS,
        id_field="specimen_uid",
    )
    if not set(specimen_ids).issubset(descriptor_rows) or not set(specimen_ids).issubset(
        provenance_rows
    ):
        raise P2FeatureError("P2 tabular sources do not cover the roster")

    scalar_rows: list[tuple[float, float, float]] = []
    for record in records:
        specimen_id = record.specimen_id
        descriptor = descriptor_rows[specimen_id]
        provenance = provenance_rows[specimen_id]
        if (
            descriptor["dataset_id"].casefold() != record.domain_id
            or provenance["source_dataset"].casefold() != record.domain_id
            or provenance["split_group"].casefold() != record.domain_id
        ):
            raise P2FeatureError(f"P2 tabular domain differs: {specimen_id}")
        if descriptor["descriptor_source"] != "ultrasonic_cscan_measurement":
            raise P2FeatureError(f"P2 descriptor source differs: {specimen_id}")
        if tuple(
            descriptor[f"{name}_unit"] for name in SCALAR_DAMAGE_NAMES
        ) != SCALAR_DAMAGE_UNITS:
            raise P2FeatureError(f"P2 scalar damage units differ: {specimen_id}")
        source_hash = _digest(
            descriptor["source_sha256"],
            label=f"physical source SHA-256 for {specimen_id}",
        )
        provenance_hash = _digest(
            provenance["cscan_source_sha256"],
            label=f"provenance C-scan SHA-256 for {specimen_id}",
        )
        if source_hash != provenance_hash:
            raise P2FeatureError(f"P2 physical provenance differs: {specimen_id}")
        expected_layup = f"{record.laminate_type}_{record.ply_count}"
        if (
            provenance["layup"].casefold() != expected_layup
            or provenance["ply_count"] != str(record.ply_count)
            or provenance["impactor"].casefold() != record.impactor
            or provenance["surface_profile_available"].casefold() != "true"
            or provenance["cscan_available"].casefold() != "true"
            or provenance["pairing_status"] != "complete_surface_cscan_cai"
            or provenance["pairing_method"] != "exact_published_specimen_id"
            or provenance["pairing_confidence"] != "exact"
        ):
            raise P2FeatureError(f"P2 specimen provenance differs: {specimen_id}")
        scalar_rows.append(
            tuple(
                _finite_csv(
                    descriptor[name], label=f"{name} for {specimen_id}"
                )
                for name in SCALAR_DAMAGE_NAMES
            )
        )

    conditions = _lvi_conditions(lvi_payload)
    total_energy: list[float] = []
    impactors: list[str] = []
    for record in records:
        condition = conditions.get(record.specimen_id)
        if condition is None:
            raise P2FeatureError(f"P2 LVI condition is missing: {record.specimen_id}")
        if (
            condition.laminate_type != record.laminate_type
            or condition.ply_count != record.ply_count
            or condition.impactor != record.impactor
        ):
            raise P2FeatureError(f"P2 LVI design differs: {record.specimen_id}")
        total_energy.append(condition.total_energy_j)
        impactors.append(condition.impactor)

    return P2FeatureAuthority(
        specimen_ids=specimen_ids,
        domain_ids=domain_ids,
        laminate_types=tuple(record.laminate_type for record in records),
        ply_counts=np.asarray([record.ply_count for record in records]),
        widths_mm=np.asarray([record.width_mm for record in records]),
        thicknesses_mm=np.asarray([record.thickness_mm for record in records]),
        surface_profile_stats=profiles,
        scalar_damage=np.asarray(scalar_rows),
        full_cscan_embedding=features[:, 0, :],
        privileged_total_energy_j=np.asarray(total_energy),
        privileged_impactors=tuple(impactors),
        full_embedding_view="FULL",
        encoder_sha256=encoder_sha256,
        embedding_state_sha256=state_sha256,
        source_sha256={
            "feature_bank": _digest(feature_bank_sha256, label="feature bank SHA-256"),
            "feature_cache": _digest(feature_cache_sha256, label="feature cache SHA-256"),
            "physical_descriptors": _digest(
                physical_descriptors_sha256,
                label="physical descriptors SHA-256",
            ),
            "provenance_specimens": _digest(
                provenance_sha256, label="specimen provenance SHA-256"
            ),
            "lvi_workbook": _digest(lvi_workbook_sha256, label="LVI workbook SHA-256"),
        },
    )


def _float_text(value: float) -> str:
    result = float(value)
    if not math.isfinite(result):
        raise P2FeatureError("feature authority contains a nonfinite value")
    return format(result, ".17g")


def serialize_feature_authority_csv(authority: P2FeatureAuthority) -> bytes:
    """Serialize feature-only provenance without embedding values or outcomes."""

    fields = (
        "specimen_id",
        "domain_id",
        "laminate_type",
        "ply_count",
        "width_mm",
        "thickness_mm",
        *PROFILE_STAT_NAMES,
        *SCALAR_DAMAGE_NAMES,
        "full_embedding_row_sha256",
        "privileged_total_impact_energy_j",
        "privileged_impactor",
    )
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(fields)
    row_hashes = authority.full_embedding_row_sha256
    for index, specimen_id in enumerate(authority.specimen_ids):
        writer.writerow(
            (
                specimen_id,
                authority.domain_ids[index],
                authority.laminate_types[index],
                str(int(authority.ply_counts[index])),
                _float_text(authority.widths_mm[index]),
                _float_text(authority.thicknesses_mm[index]),
                *(
                    _float_text(value)
                    for value in authority.surface_profile_stats[index]
                ),
                *(_float_text(value) for value in authority.scalar_damage[index]),
                row_hashes[index],
                _float_text(authority.privileged_total_energy_j[index]),
                authority.privileged_impactors[index],
            )
        )
    return buffer.getvalue().encode("ascii")
