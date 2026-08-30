from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from collections import Counter
from collections.abc import Iterable, Mapping
from collections.abc import Set as AbstractSet
from dataclasses import dataclass
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook

from cmc_bbdm.damage_response.authority import (
    AuthorityError,
    FileSnapshot,
    snapshot_file,
)
from cmc_bbdm.damage_response.contracts import PRIMARY_COUNTS
from cmc_bbdm.damage_response.pairing import FeatureIdentity
from cmc_bbdm.damage_response.targets import (
    PublishedPeak,
    decimal_places_from_excel_format,
)

_SHA256_RE = re.compile(r"[0-9a-f]{64}")
_INVENTORY_KEYS = {
    "dataset_id",
    "file_id",
    "filename",
    "folder",
    "relative_path",
    "sha256",
    "size",
    "url",
    "version",
}
OFFICIAL_FOLDER_COUNTS: Mapping[str, int] = {
    "1_Low velocity impact testing condition": 1,
    "2_Specimen size": 1,
    "3_Specimen image": 892,
    "4_Compression after impact testing raw data": 446,
    "5_Compression after impact strength": 1,
}
LAMINATE_CATEGORIES = ("cross_ply", "quasi_isotropic")
IMPACTOR_CATEGORIES = ("coni120", "coni60", "flat", "hemia", "hemib", "hemic")


class SourceError(RuntimeError):
    """Raised when official or historical source metadata is inconsistent."""


@dataclass(frozen=True)
class OfficialFileRecord:
    dataset_id: str
    file_id: str
    filename: str
    folder: str
    relative_path: str
    sha256: str
    size: int
    version: int


def _expected_digest(value: str, label: str) -> str:
    if not isinstance(value, str):
        raise SourceError(f"{label} must be a SHA-256")
    digest = value.strip().casefold()
    if _SHA256_RE.fullmatch(digest) is None:
        raise SourceError(f"{label} must be a SHA-256")
    return digest


def _bound_snapshot(
    path: Path,
    *,
    expected_sha256: str,
    logical_source: str,
    relative_path: str,
    max_bytes: int,
) -> FileSnapshot:
    expected = _expected_digest(expected_sha256, f"{logical_source} SHA-256")
    try:
        snapshot = snapshot_file(
            path,
            max_bytes=max_bytes,
            logical_source=logical_source,
            relative_path=relative_path,
        )
    except AuthorityError as error:
        raise SourceError(str(error)) from error
    if snapshot.sha256 != expected:
        raise SourceError(
            f"{logical_source} SHA-256 mismatch: "
            f"expected {expected}, observed {snapshot.sha256}"
        )
    return snapshot


def load_official_inventory(
    path: Path,
    *,
    expected_sha256: str,
    expected_folder_counts: Mapping[str, int] = OFFICIAL_FOLDER_COUNTS,
) -> tuple[OfficialFileRecord, ...]:
    """Load the exact Mendeley v3 inventory without retaining download URLs."""

    snapshot = _bound_snapshot(
        path,
        expected_sha256=expected_sha256,
        logical_source="official inventory",
        relative_path="hasebe_cai/inventory.jsonl",
        max_bytes=32 * 1024 * 1024,
    )
    try:
        payload = Path(path).read_bytes()
    except OSError as error:
        raise SourceError("official inventory cannot be read") from error
    if len(payload) != snapshot.size or hashlib.sha256(payload).hexdigest() != snapshot.sha256:
        raise SourceError("official inventory changed during read")

    records: list[OfficialFileRecord] = []
    file_ids: set[str] = set()
    relative_paths: set[str] = set()
    for line_number, line in enumerate(payload.decode("utf-8").splitlines(), start=1):
        try:
            value = json.loads(line)
        except json.JSONDecodeError as error:
            raise SourceError(f"invalid inventory JSON at line {line_number}") from error
        if not isinstance(value, dict) or set(value) != _INVENTORY_KEYS:
            raise SourceError(f"official inventory schema drift at line {line_number}")
        dataset_id = value["dataset_id"]
        version = value["version"]
        file_id = value["file_id"]
        filename = value["filename"]
        folder = value["folder"]
        relative_path = value["relative_path"]
        digest = value["sha256"]
        size = value["size"]
        url = value["url"]
        if dataset_id != "8scdmfdcfb" or version != 3:
            raise SourceError("official inventory dataset/version identity changed")
        if not isinstance(file_id, str) or not file_id or file_id in file_ids:
            raise SourceError("official inventory file IDs are empty or duplicate")
        if (
            not isinstance(filename, str)
            or not filename
            or PurePosixPath(filename).name != filename
        ):
            raise SourceError("official inventory filename is unsafe")
        if not isinstance(folder, str) or folder not in expected_folder_counts:
            raise SourceError(f"unexpected official inventory folder: {folder!r}")
        if not isinstance(relative_path, str):
            raise SourceError("official inventory relative path is invalid")
        candidate = PurePosixPath(relative_path)
        expected_path = PurePosixPath("8scdmfdcfb") / "v3" / folder / filename
        if (
            candidate.is_absolute()
            or ".." in candidate.parts
            or candidate != expected_path
            or relative_path in relative_paths
        ):
            raise SourceError("official inventory relative path is unsafe or duplicate")
        if not isinstance(digest, str) or _SHA256_RE.fullmatch(digest) is None:
            raise SourceError("official inventory file SHA-256 is invalid")
        if not isinstance(size, int) or isinstance(size, bool) or size <= 0:
            raise SourceError("official inventory file size is invalid")
        if not isinstance(url, str) or not url.startswith("https://data.mendeley.com/"):
            raise SourceError("official inventory download URL is invalid")
        file_ids.add(file_id)
        relative_paths.add(relative_path)
        records.append(
            OfficialFileRecord(
                dataset_id=dataset_id,
                file_id=file_id,
                filename=filename,
                folder=folder,
                relative_path=relative_path,
                sha256=digest,
                size=size,
                version=version,
            )
        )

    observed_counts = Counter(record.folder for record in records)
    if dict(observed_counts) != dict(expected_folder_counts):
        raise SourceError(
            "official inventory folder counts differ: "
            f"expected {dict(expected_folder_counts)!r}, observed {dict(observed_counts)!r}"
        )
    return tuple(sorted(records, key=lambda record: record.relative_path))


@dataclass(frozen=True)
class SpecimenSize:
    specimen_id: str
    height_mm: float
    width_mm: float
    thickness_mm: float


@dataclass(frozen=True)
class PreCaiObservation:
    specimen_id: str
    has_numeric_damage_observation: bool
    is_intact: bool
    included: bool


@dataclass(frozen=True)
class DesignMetadata:
    specimen_id: str
    domain_id: str
    laminate_type: str
    ply_count: int
    impactor: str
    width_mm: float
    thickness_mm: float


def _normal_specimen_id(value: object, *, row: int) -> str:
    if not isinstance(value, str) or not value.strip():
        raise SourceError(f"empty specimen ID at workbook row {row}")
    return value.strip().casefold()


def _positive_number(value: object, *, label: str, row: int) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise SourceError(f"nonnumeric {label} at workbook row {row}")
    result = float(value)
    if not math.isfinite(result) or result <= 0.0:
        raise SourceError(f"invalid {label} at workbook row {row}")
    return result


def _open_bound_workbook(path: Path, *, expected_sha256: str, relative_path: str):
    _bound_snapshot(
        path,
        expected_sha256=expected_sha256,
        logical_source=f"official workbook {Path(relative_path).name}",
        relative_path=relative_path,
        max_bytes=128 * 1024 * 1024,
    )
    try:
        return load_workbook(path, data_only=True, read_only=False)
    except (OSError, ValueError) as error:
        raise SourceError(f"official workbook cannot be decoded: {Path(relative_path).name}") from error


def _require_headers(sheet, expected: Mapping[str, object]) -> None:
    changed = {
        address: (sheet[address].value, expected_value)
        for address, expected_value in expected.items()
        if sheet[address].value != expected_value
    }
    if changed:
        raise SourceError(f"registered workbook headers changed: {changed}")


def read_specimen_sizes(
    path: Path, *, expected_sha256: str
) -> dict[str, SpecimenSize]:
    workbook = _open_bound_workbook(
        path,
        expected_sha256=expected_sha256,
        relative_path="8scdmfdcfb/v3/2_Specimen size/Specimen size_v0.2.xlsx",
    )
    try:
        if "Specimen size" not in workbook.sheetnames:
            raise SourceError("registered Specimen size worksheet is absent")
        sheet = workbook["Specimen size"]
        _require_headers(
            sheet,
            {
                "B2": "Specimen No.",
                "C2": "Specimen size[mm]",
                "C3": "Height",
                "D3": "Width",
                "E3": "Thickness",
            },
        )
        records: dict[str, SpecimenSize] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=4, min_col=2, max_col=5, values_only=True),
            start=4,
        ):
            if row[0] is None:
                continue
            specimen_id = _normal_specimen_id(row[0], row=row_number)
            if specimen_id in records:
                raise SourceError(f"duplicate specimen size ID: {specimen_id}")
            numeric = tuple(
                isinstance(value, (int, float)) and not isinstance(value, bool)
                for value in row[1:]
            )
            if numeric == (False, False, False):
                continue
            if numeric != (True, True, True):
                raise SourceError(f"partially numeric specimen size at row {row_number}")
            records[specimen_id] = SpecimenSize(
                specimen_id=specimen_id,
                height_mm=_positive_number(row[1], label="height", row=row_number),
                width_mm=_positive_number(row[2], label="width", row=row_number),
                thickness_mm=_positive_number(
                    row[3], label="thickness", row=row_number
                ),
            )
    finally:
        workbook.close()
    if not records:
        raise SourceError("specimen size workbook contains no records")
    return records


def read_published_peaks(
    path: Path, *, expected_sha256: str
) -> dict[str, PublishedPeak]:
    workbook = _open_bound_workbook(
        path,
        expected_sha256=expected_sha256,
        relative_path=(
            "8scdmfdcfb/v3/5_Compression after impact strength/"
            "Compression after impact strength_v0.2.xlsx"
        ),
    )
    try:
        if "CAI strength" not in workbook.sheetnames:
            raise SourceError("registered CAI strength worksheet is absent")
        sheet = workbook["CAI strength"]
        _require_headers(
            sheet,
            {
                "B2": "Specimen No.",
                "C2": "Compression after impact strength",
                "C3": "[Mpa]",
                "D3": "[%]",
            },
        )
        records: dict[str, PublishedPeak] = {}
        for row_number in range(4, sheet.max_row + 1):
            identity = sheet.cell(row_number, 2).value
            if identity is None:
                continue
            specimen_id = _normal_specimen_id(identity, row=row_number)
            if specimen_id in records:
                raise SourceError(f"duplicate published peak ID: {specimen_id}")
            cell = sheet.cell(row_number, 3)
            if not isinstance(cell.value, (int, float)) or isinstance(cell.value, bool):
                continue
            records[specimen_id] = PublishedPeak(
                specimen_id=specimen_id,
                value_mpa=_positive_number(
                    cell.value, label="published CAI strength", row=row_number
                ),
                decimal_places=decimal_places_from_excel_format(cell.number_format),
            )
    finally:
        workbook.close()
    if not records:
        raise SourceError("published peak workbook contains no numeric records")
    return records


def read_lvi_observations(
    path: Path, *, expected_sha256: str
) -> dict[str, PreCaiObservation]:
    workbook = _open_bound_workbook(
        path,
        expected_sha256=expected_sha256,
        relative_path=(
            "8scdmfdcfb/v3/1_Low velocity impact testing condition/"
            "Low velocity impact testing condition and damage_v0.2.xlsx"
        ),
    )
    try:
        if "LVI condition" not in workbook.sheetnames:
            raise SourceError("registered LVI condition worksheet is absent")
        sheet = workbook["LVI condition"]
        _require_headers(
            sheet,
            {
                "B2": "Specimen No.",
                "C3": "Layup",
                "D3": "Impactor shape",
                "E3": "Impact energy",
                "G3": "Projected delamination area",
                "H3": "Dent depth",
                "I3": "Is included",
                "E4": "[J/mm]",
                "F4": "[J]",
            },
        )
        records: dict[str, PreCaiObservation] = {}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=5, min_col=2, max_col=9, values_only=True),
            start=5,
        ):
            if row[0] is None:
                continue
            specimen_id = _normal_specimen_id(row[0], row=row_number)
            if specimen_id in records:
                raise SourceError(f"duplicate LVI observation ID: {specimen_id}")
            area_numeric = isinstance(row[5], (int, float)) and not isinstance(
                row[5], bool
            )
            dent_numeric = isinstance(row[6], (int, float)) and not isinstance(
                row[6], bool
            )
            if area_numeric != dent_numeric:
                raise SourceError(
                    f"partially numeric LVI damage observation at row {row_number}"
                )
            records[specimen_id] = PreCaiObservation(
                specimen_id=specimen_id,
                has_numeric_damage_observation=area_numeric and dent_numeric,
                is_intact=(str(row[2]).strip() == "-" and not area_numeric),
                included=str(row[7]).strip().casefold() == "yes",
            )
    finally:
        workbook.close()
    if not records:
        raise SourceError("LVI workbook contains no observations")
    return records


def read_primary_design_metadata(
    path: Path,
    expected_sha256: str,
    primary_identities: Iterable[FeatureIdentity],
    specimen_sizes: Mapping[str, SpecimenSize],
) -> tuple[DesignMetadata, ...]:
    """Bind path-free design metadata to the frozen 276-specimen roster."""

    snapshot = _bound_snapshot(
        path,
        expected_sha256=expected_sha256,
        logical_source="legacy spatial-pair design metadata",
        relative_path="hasebe/manifest/paired.csv",
        max_bytes=32 * 1024 * 1024,
    )
    try:
        payload = Path(path).read_bytes()
    except OSError as error:
        raise SourceError("legacy design metadata cannot be read") from error
    if (
        len(payload) != snapshot.size
        or hashlib.sha256(payload).hexdigest() != snapshot.sha256
    ):
        raise SourceError("legacy design metadata changed during read")

    identities = tuple(primary_identities)
    expected_count = sum(PRIMARY_COUNTS.values())
    observed_counts = Counter(identity.domain_id for identity in identities)
    specimen_ids = tuple(identity.specimen_id for identity in identities)
    if (
        len(identities) != expected_count
        or len(set(specimen_ids)) != expected_count
        or dict(observed_counts) != dict(PRIMARY_COUNTS)
    ):
        raise SourceError("design metadata requires the frozen primary roster")

    try:
        reader = csv.DictReader(payload.decode("utf-8-sig").splitlines())
    except UnicodeDecodeError as error:
        raise SourceError("legacy design metadata is not UTF-8") from error
    required = {
        "sample_id",
        "dataset_id",
        "laminate_type",
        "ply_count",
        "impactor",
    }
    if reader.fieldnames is None or not required.issubset(reader.fieldnames):
        raise SourceError("legacy design metadata schema drift")

    manifest_rows: dict[str, tuple[str, str, int, str]] = {}
    for row_number, row in enumerate(reader, start=2):
        specimen_id = _normal_specimen_id(row["sample_id"], row=row_number)
        if specimen_id in manifest_rows:
            raise SourceError(f"duplicate legacy design specimen ID: {specimen_id}")
        domain_id = str(row["dataset_id"] or "").strip().casefold()
        laminate_type = str(row["laminate_type"] or "").strip().casefold()
        impactor = str(row["impactor"] or "").strip().casefold()
        try:
            ply_count = int(str(row["ply_count"] or "").strip())
        except ValueError as error:
            raise SourceError(
                f"invalid legacy design ply count at row {row_number}"
            ) from error
        if not domain_id:
            raise SourceError(f"empty legacy design domain at row {row_number}")
        if laminate_type not in LAMINATE_CATEGORIES:
            raise SourceError(
                f"unknown legacy laminate category at row {row_number}"
            )
        if impactor not in IMPACTOR_CATEGORIES:
            raise SourceError(f"unknown legacy impactor category at row {row_number}")
        if ply_count <= 0:
            raise SourceError(f"invalid legacy design ply count at row {row_number}")
        manifest_rows[specimen_id] = (
            domain_id,
            laminate_type,
            ply_count,
            impactor,
        )

    normalized_sizes = {
        specimen_id.strip().casefold(): size
        for specimen_id, size in specimen_sizes.items()
    }
    records: list[DesignMetadata] = []
    for identity in identities:
        specimen_id = identity.specimen_id.strip().casefold()
        domain_id = identity.domain_id.strip().casefold()
        manifest = manifest_rows.get(specimen_id)
        if manifest is None:
            raise SourceError(f"primary design metadata is missing: {specimen_id}")
        if manifest[0] != domain_id:
            raise SourceError(f"primary design domain differs: {specimen_id}")
        size = normalized_sizes.get(specimen_id)
        if size is None:
            raise SourceError(f"primary specimen size is missing: {specimen_id}")
        width = float(size.width_mm)
        thickness = float(size.thickness_mm)
        if not math.isfinite(width) or width <= 0.0:
            raise SourceError(f"primary specimen width is invalid: {specimen_id}")
        if not math.isfinite(thickness) or thickness <= 0.0:
            raise SourceError(f"primary specimen thickness is invalid: {specimen_id}")
        records.append(
            DesignMetadata(
                specimen_id=specimen_id,
                domain_id=domain_id,
                laminate_type=manifest[1],
                ply_count=manifest[2],
                impactor=manifest[3],
                width_mm=width,
                thickness_mm=thickness,
            )
        )
    return tuple(records)


@dataclass(frozen=True)
class SpatialPairIdentity:
    specimen_id: str
    domain_id: str
    surface_image_sha256: str
    cscan_sha256: str
    surface_profile_available: bool


@dataclass(frozen=True)
class SpatialExpansion:
    identity_pair_ids: tuple[str, ...]
    valid_trace_pair_ids: tuple[str, ...]
    extra_identity_pair_ids: tuple[str, ...]
    extra_valid_trace_pair_ids: tuple[str, ...]


def classify_spatial_expansion(
    *,
    raw_identity_ids: AbstractSet[str],
    valid_trace_ids: AbstractSet[str],
    spatial_ids: AbstractSet[str],
    primary_ids: AbstractSet[str],
) -> SpatialExpansion:
    """Separate exact file identity intersections from usable decoded traces."""

    raw = {value.strip().casefold() for value in raw_identity_ids}
    valid = {value.strip().casefold() for value in valid_trace_ids}
    spatial = {value.strip().casefold() for value in spatial_ids}
    primary = {value.strip().casefold() for value in primary_ids}
    if "" in raw | valid | spatial | primary:
        raise SourceError("spatial expansion identities must be nonempty")
    if not valid.issubset(raw):
        raise SourceError("valid trace identities are not a subset of raw identities")
    identity_pairs = raw & spatial
    valid_pairs = valid & spatial
    return SpatialExpansion(
        identity_pair_ids=tuple(sorted(identity_pairs)),
        valid_trace_pair_ids=tuple(sorted(valid_pairs)),
        extra_identity_pair_ids=tuple(sorted(identity_pairs - primary)),
        extra_valid_trace_pair_ids=tuple(sorted(valid_pairs - primary)),
    )


def load_spatial_pairs(
    path: Path, *, expected_sha256: str
) -> tuple[SpatialPairIdentity, ...]:
    snapshot = _bound_snapshot(
        path,
        expected_sha256=expected_sha256,
        logical_source="legacy spatial-pair manifest",
        relative_path="hasebe/manifest/paired.csv",
        max_bytes=32 * 1024 * 1024,
    )
    payload = Path(path).read_bytes()
    if len(payload) != snapshot.size or hashlib.sha256(payload).hexdigest() != snapshot.sha256:
        raise SourceError("legacy spatial-pair manifest changed during read")
    try:
        reader = csv.DictReader(payload.decode("utf-8-sig").splitlines())
        required = {
            "sample_id",
            "dataset_id",
            "source_sha256",
            "target_sha256",
            "surface_profile_before_path",
            "surface_profile_after_path",
        }
        if reader.fieldnames is None or not required.issubset(reader.fieldnames):
            raise SourceError("legacy spatial-pair manifest schema drift")
        records: list[SpatialPairIdentity] = []
        seen: set[str] = set()
        for row_number, row in enumerate(reader, start=2):
            specimen_id = _normal_specimen_id(row["sample_id"], row=row_number)
            domain_id = str(row["dataset_id"] or "").strip().casefold()
            source_digest = str(row["source_sha256"] or "").strip().casefold()
            target_digest = str(row["target_sha256"] or "").strip().casefold()
            if specimen_id in seen:
                raise SourceError(f"duplicate legacy spatial specimen ID: {specimen_id}")
            if not domain_id:
                raise SourceError(f"empty legacy spatial domain at row {row_number}")
            if (
                _SHA256_RE.fullmatch(source_digest) is None
                or _SHA256_RE.fullmatch(target_digest) is None
            ):
                raise SourceError(f"invalid legacy spatial SHA-256 at row {row_number}")
            seen.add(specimen_id)
            records.append(
                SpatialPairIdentity(
                    specimen_id=specimen_id,
                    domain_id=domain_id,
                    surface_image_sha256=source_digest,
                    cscan_sha256=target_digest,
                    surface_profile_available=bool(
                        str(row["surface_profile_before_path"] or "").strip()
                        and str(row["surface_profile_after_path"] or "").strip()
                    ),
                )
            )
    except UnicodeDecodeError as error:
        raise SourceError("legacy spatial-pair manifest is not UTF-8") from error
    return tuple(sorted(records, key=lambda record: record.specimen_id))
